import requests
from bs4 import BeautifulSoup
import openpyxl

excel = openpyxl.Workbook()
print(excel.sheetnames)
sheet = excel.active
sheet.title = 'top rated movie by ambar'
print(excel.sheetnames)
sheet.append(['movie rank, movie name, release year, imbd rating'])


source = requests.get('https://www.imdb.com/chart/top/')
source.raise_for_status()

soup = BeautifulSoup(source.text, 'html.parser')
#print(soup)
movies = soup.find('tbody', class_="lister-list").find_all('tr')
#print(movies)

for movie in movies:

    name = movie.find('td', class_="titleColumn").a.text

    rank = movie.find('td', class_="titleColumn").get_text(strip=True).split('.')[0]

    year = movie.find('td', class_="titleColumn").span.text.strip('()')

    rating = movie.find('td', class_="ratingColumn imdbRating").text

    print(rank, name, year, rating)  
    
    sheet.append([rank, name, year, rating])

excel.save('movie rating by ambar.xlsx')
